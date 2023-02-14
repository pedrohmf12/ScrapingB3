# Imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors
from openpyxl.styles import Font, Border, Side, PatternFill, Font, GradientFill, Alignment
from time import sleep
import win32com.client as win32
import win32com
import win32gui
from PIL import ImageGrab, Image
from datetime import datetime
from babel.dates import format_datetime


options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--disable-gpu')

driver = webdriver.Chrome(options=options)

def get_key(dictionary, value):
    for key in dictionary:
        if dictionary[key] == value:
            return key


ativos = {
    "Janeiro": "F",
    "Fevereiro": "G",
    "Março": "H",
    "Abril": "J",
    "Maio": "K",
    "Junho": "M",
    "Julho": "N",
    "Agosto": "Q",
    "Setembro": "U",
    "Outubro": "V",
    "Novembro": "X",
    "Dezembro": "Z"
}

mes_atual_datetime = datetime.now()
formatted_date = format_datetime(mes_atual_datetime, format='MMMM', locale='pt_BR')
mes_atual = formatted_date
mes_atual = mes_atual.capitalize()
mes_atual_abreviado = ativos[mes_atual]


meses = list(ativos.keys())
indice_mes_atual = meses.index(mes_atual)

meses_subjacentes = meses[indice_mes_atual:indice_mes_atual + 5]
ativos_mes=[]
for mes in meses_subjacentes:
    #print("Mês:", mes, "Abreviatura:", ativos[mes])
    ativos_mes.append(ativos[mes])

preco_ativo=[]
for ativo in ativos_mes:
    driver.get(f"https://br.tradingview.com/symbols/BMFBOVESPA-BGI{ativo}2023/")
    sleep(2)
    cotacao = driver.find_element(By.XPATH,'//*[@id="anchor-page-1"]/div/div[3]/div[1]/div/div/div/div[1]/div[1]').text
    var_preco = driver.find_element(By.XPATH,'//*[@id="anchor-page-1"]/div/div[3]/div[1]/div/div/div/div[1]/div[3]/span[1]').text
    var_perc = driver.find_element(By.XPATH,'//*[@id="anchor-page-1"]/div/div[3]/div[1]/div/div/div/div[1]/div[3]/span[2]').text

    ativos_values = [cotacao, var_preco, var_perc]
    preco_ativo.append(ativos_values)


df = pd.DataFrame(columns=['MÊS','PREÇO', 'VAR.R$', 'VAR.%'])

meses_cotados = []
for cada in ativos_mes:
    value = f'{cada}'
    key = get_key(ativos, value)
    key = key[0:3].upper()
    meses_cotados.append(key)
    print(meses_cotados)

i = 0
for precos in preco_ativo:
    #print(precos)
    data = {'MÊS': [meses_cotados[i]],
            'PREÇO': [precos[0]], 
            'VAR.R$': [precos[1]], 
            'VAR.%': [precos[2]]}
    i= i + 1
    df = pd.concat([df, pd.DataFrame(data)], ignore_index=True)
    

planilha = df.to_excel('Arroba_Boi_Futuro.xlsx', index=False)

#importando DataFrama para excel
file_name = 'Arroba_Boi_Futuro.xlsx'
wb = openpyxl.load_workbook(file_name)
ws= wb.active

#Parâmetros de Preenchimento
my_fill = PatternFill(start_color='eebe2c',
                   end_color='eebe2c',
                   fill_type='solid')
my_fill_zebra = PatternFill(start_color='f6eab5',
                            end_color='f6eab5',
                            fill_type='solid')

#Parâmetros da Fonte
my_font = Font(bold=True)

#Parâmetros da Borda
borders = Border(left=Side(border_style="thin", color='000000'))

# Formata o cabeçalho
my_header = ['A1', 'B1', 'C1', 'D1']
for cell in my_header:
    ws[cell].fill = my_fill
    ws[cell].border = borders
    ws[cell].font = my_font
    ws[cell].alignment = Alignment(horizontal= 'center', vertical='center')

ws.sheet_view.showGridLines = False

price_values =['B2','B3','B4','B5','B6'] 
price_cotation = []
for i in price_values: 
    valores = ws[f'{i}'].value
    valores_float = float(valores)
    price_cotation.append(valores_float)

variations = ['C2','C3','C4','C5','C6'] 
variations_price = []
for i in variations: 
    valores = ws[f'{i}'].value
    fixed_string = valores.replace('+','').replace("−", "-")
    float_value = float(fixed_string)
    variations_price.append(float_value)

variations_perc = ['D2','D3','D4','D5','D6'] 
variation_perc = []
for i in variations_perc: 
    valores = ws[f'{i}'].value
    fixed_string = valores.replace('(','').replace(')','').replace('+','')
    #float_value = float(fixed_string)
    variation_perc.append(fixed_string)


cells = ['A','B','C','D']
bordas = ['B','C','D']
moedas = ['B','C']


for i in range(5):
    ws[f'B{i+2}'] = price_cotation[i]
    ws[f'C{i+2}'] = variations_price[i]
    ws[f'D{i+2}'] = variation_perc[i]
    #ws[f'D{i+2}'].number_format = '0.00%'
    for cell in cells:
        ws[f'{cell}'f'{i+2}'].alignment = Alignment(horizontal= 'center', vertical='center')
    for cell in cells:
        if i%2 == 0:
            ws[f'{cell}'f'{i+2}'].fill = my_fill_zebra
        else:
            ws[f'{cell}'f'{i+2}'].fill = my_fill
        #print(i)
    for cell in bordas:
        ws[f'{cell}'f'{i+2}'].border = borders
    for cell in moedas:
        ws[f'{cell}'f'{i+2}'].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
#print('exito')

ws.row_dimensions[1].height = 40


wb.save(file_name)

# Obtenha o diretório atual
current_dir = os.getcwd()

# Abra a planilha do Excel
excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
excel.Visible = False

file_path = os.path.join(current_dir, file_name)
#print(file_path)
if os.path.exists(file_path):
    workbook = excel.Workbooks.Open(file_path)
    #print(workbook)
    sheet = workbook.ActiveSheet
    sheet.Range("A1:D6").CopyPicture(Format=win32.constants.xlBitmap)
else:
    print("O arquivo '{}' não foi encontrado".format(file_name))

# Capturar a imagem na área de transferência
image = ImageGrab.grabclipboard()

# Defina o caminho para salvar a imagem
image_path = os.path.join(os.getcwd(),'planilha.bmp')

# Salve a imagem em alta resolução
image.save(image_path, quality=100)

image = Image.open("planilha.bmp")

# Salve a imagem como um arquivo PNG
image.save(".\imagens\cotacao.png", "PNG")

os.remove("planilha.bmp")

# Feche a planilha
workbook.Close()
excel.Quit()

